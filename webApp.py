from flask import Flask, request, jsonify, send_from_directory , url_for
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import requests
import os
from datetime import datetime
from PIL import Image as PILImage


application = Flask(__name__)

# Directory to store exported files
EXPORT_FOLDER = 'exported_files'
os.makedirs(EXPORT_FOLDER, exist_ok=True)

TMP_IMAGE_FOLDER = 'tmp_images'
os.makedirs(TMP_IMAGE_FOLDER, exist_ok=True)

@application.route('/')
def index():
    return "Hello Ashish"

@application.route('/upload', methods=['POST'])
def upload():
    data = request.json
    json_data = data['data']
    file_name = data.get('fileName')

    if not file_name:
        file_name = datetime.now().strftime("%d-%m-%Y-%M-%S") + '_export_' + '.xlsx'

    # Save files in the specified folder
    file_path = os.path.join(EXPORT_FOLDER, file_name)

    # Extract headers from JSON data
    headers = list(json_data[0].keys())

    # Create a Pandas DataFrame from the JSON data
    df = pd.DataFrame(json_data)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    # Iterate through the DataFrame and add data to the worksheet
    for index, row in df.iterrows():
        for col_num, header in enumerate(headers, start=1):
            if header != 'image':
                ws.cell(row=index + 2, column=col_num, value=row[header])
            else:
                # Add image to the worksheet
                response = requests.get(row['image'])
                img = PILImage.open(BytesIO(response.content))
                local_image_path = TMP_IMAGE_FOLDER+'/local_image'+ datetime.now().strftime("%d-%m-%Y-%M-%S") + '.jpg'
                img.save(local_image_path)
                img = Image(local_image_path)


                # img = Image(BytesIO(response.content))
                img.width = 80  # Adjust the width as needed
                img.height = 80  # Adjust the height as needed

                col_idx = headers.index('image')
                row_idx = index + 2

                ws.row_dimensions[row_idx].height = img.height + 2
                ws.column_dimensions[chr(ord("A") + col_idx)].width = img.width / 7

                ws.add_image(img, f'{chr(ord("A") + col_idx)}{row_idx}')

    # Save the workbook in the designated folder
    wb.save(file_path)
    full_url = request.host_url.rstrip('/') + url_for('download', filename=file_name)

    # return jsonify({'message': 'File created', 'fileName': file_name, 'downloadLink': f'/download/{file_name}'})
    return jsonify({'message': 'File created', 'fileName': file_name, 'downloadLink': full_url})

@application.route('/download/<filename>')
def download(filename):
    # Updated to serve files from the EXPORT_FOLDER
    return send_from_directory(directory=EXPORT_FOLDER, path=filename, as_attachment=True)

if __name__ == '__main__':
    application.run(debug=True)
